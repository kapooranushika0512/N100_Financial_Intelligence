import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DB_PATH = "db/nifty100.db"
OUTPUT_FILE = "output/peer_comparison.xlsx"


GREEN = PatternFill(fill_type="solid", start_color="C6EFCE")
YELLOW = PatternFill(fill_type="solid", start_color="FFF2CC")
RED = PatternFill(fill_type="solid", start_color="F4CCCC")
GOLD = PatternFill(fill_type="solid", start_color="FFD966")


def load_data():
    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql("SELECT * FROM financial_ratios", conn)
    market = pd.read_sql("SELECT * FROM market_cap", conn)
    analysis = pd.read_sql("SELECT * FROM analysis", conn)
    peer_groups = pd.read_sql("SELECT * FROM peer_groups", conn)
    peer_percentiles = pd.read_sql("SELECT * FROM peer_percentiles", conn)

    conn.close()

    if "id" in market.columns:
        market = market.drop(columns=["id"])

    if "id" in analysis.columns:
        analysis = analysis.drop(columns=["id"])

    ratios["year"] = ratios["year"].astype(str)

    if "year" in market.columns:
        market["year"] = market["year"].astype(str)

    if "year" in peer_percentiles.columns:
        peer_percentiles["year"] = peer_percentiles["year"].astype(str)

    df = ratios.merge(
        peer_groups,
        on="company_id",
        how="left",
        suffixes=("", "_peer")
    )

    if "year" in market.columns:
        df = df.merge(
            market,
            on=["company_id", "year"],
            how="left"
        )
    else:
        df = df.merge(
            market,
            on="company_id",
            how="left"
        )

    df = df.merge(
        analysis,
        on="company_id",
        how="left"
    )

    return df, peer_percentiles


def export_excel(df, peer_percentiles, limit=None):

    os.makedirs("output", exist_ok=True)

    writer = pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    )

    groups = sorted(df["peer_group_name"].dropna().unique())

    if limit:
        groups = groups[:limit]

    for group in groups:

        sheet = df[df["peer_group_name"] == group].copy()

        sheet = sheet.merge(
            peer_percentiles[
                ["company_id", "year", "metric", "percentile_rank"]
            ],
            on=["company_id", "year"],
            how="left"
        )

        sheet.to_excel(
            writer,
            sheet_name=group[:31],
            index=False
        )

    writer.close()


def apply_formatting():

    wb = load_workbook(OUTPUT_FILE)

    for ws in wb.worksheets:

        headers = [c.value for c in ws[1]]

        benchmark_col = None

        if "is_benchmark" in headers:
            benchmark_col = headers.index("is_benchmark") + 1

        percentile_cols = []

        for i, h in enumerate(headers, start=1):
            if "percentile" in str(h).lower():
                percentile_cols.append(i)

        for row in range(2, ws.max_row + 1):

            if benchmark_col:

                if ws.cell(row=row, column=benchmark_col).value == 1:

                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = GOLD

            for col in percentile_cols:

                value = ws.cell(row=row, column=col).value

                if value is None:
                    continue

                if value >= 0.75:
                    ws.cell(row=row, column=col).fill = GREEN
                elif value <= 0.25:
                    ws.cell(row=row, column=col).fill = RED
                else:
                    ws.cell(row=row, column=col).fill = YELLOW

    wb.save(OUTPUT_FILE)


def run(limit=None):

    df, peer_percentiles = load_data()

    export_excel(df, peer_percentiles, limit)

    apply_formatting()

    print("Peer comparison report generated.")

    return df


if __name__ == "__main__":
    run()